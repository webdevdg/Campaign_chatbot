import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule

from services.data_service import load_data

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")

# ── Color palette ──
NAVY = "1B2A4A"
DARK_BLUE = "2C3E6B"
ACCENT_BLUE = "4472C4"
ACCENT_GREEN = "548235"
ACCENT_RED = "C0392B"
LIGHT_GRAY = "F2F2F2"
MID_GRAY = "D9D9D9"
WHITE = "FFFFFF"
SECTION_BG = "E8EDF5"

# ── Reusable styles ──
thin_border = Border(
    left=Side(style="thin", color=MID_GRAY),
    right=Side(style="thin", color=MID_GRAY),
    top=Side(style="thin", color=MID_GRAY),
    bottom=Side(style="thin", color=MID_GRAY),
)

header_font = Font(bold=True, color=WHITE, size=10, name="Calibri")
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

subheader_font = Font(bold=True, color=NAVY, size=10, name="Calibri")
subheader_fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")

title_font = Font(bold=True, color=NAVY, size=20, name="Calibri")
subtitle_font = Font(bold=False, color="666666", size=11, name="Calibri")
kpi_label_font = Font(bold=True, color="555555", size=9, name="Calibri")
kpi_value_font = Font(bold=True, color=NAVY, size=16, name="Calibri")
kpi_sub_font = Font(bold=False, color="888888", size=8, name="Calibri")

body_font = Font(size=10, name="Calibri")
bold_body = Font(bold=True, size=10, name="Calibri", color=NAVY)

stripe_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
green_font = Font(color=ACCENT_GREEN, bold=True, size=10, name="Calibri")
red_font = Font(color=ACCENT_RED, bold=True, size=10, name="Calibri")


def _auto_width(ws, min_width=10, max_width=28):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        longest = max((len(str(cell.value or "")) for cell in col), default=min_width)
        ws.column_dimensions[col_letter].width = min(max(longest + 3, min_width), max_width)


def _write_table(ws, df, start_row, start_col=1, currency_cols=None, pct_cols=None, ratio_cols=None, int_cols=None):
    """Write a DataFrame as a formatted table starting at (start_row, start_col)."""
    currency_cols = currency_cols or set()
    pct_cols = pct_cols or set()
    ratio_cols = ratio_cols or set()
    int_cols = int_cols or set()

    headers = list(df.columns)
    for ci, h in enumerate(headers, start_col):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for ri, row in enumerate(df.itertuples(index=False), start_row + 1):
        is_stripe = (ri - start_row) % 2 == 0
        for ci, (col_name, val) in enumerate(zip(headers, row), start_col):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_stripe:
                cell.fill = stripe_fill
            if col_name in currency_cols:
                cell.number_format = '$#,##0.00'
            elif col_name in pct_cols:
                cell.number_format = '0.00%'
            elif col_name in ratio_cols:
                cell.number_format = '0.00x'
            elif col_name in int_cols:
                cell.number_format = '#,##0'

    return start_row + len(df) + 1


def _kpi_card(ws, row, col, label, value, sub_text=""):
    """Write a KPI card: label on top, large value, optional sub-text."""
    cell_label = ws.cell(row=row, column=col, value=label.upper())
    cell_label.font = kpi_label_font
    cell_label.alignment = Alignment(horizontal="center")

    cell_val = ws.cell(row=row + 1, column=col, value=value)
    cell_val.font = kpi_value_font
    cell_val.alignment = Alignment(horizontal="center")

    if sub_text:
        cell_sub = ws.cell(row=row + 2, column=col, value=sub_text)
        cell_sub.font = kpi_sub_font
        cell_sub.alignment = Alignment(horizontal="center")


def _apply_filter(df, parameters):
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"])

    if parameters.get("brand"):
        brands = parameters["brand"] if isinstance(parameters["brand"], list) else [parameters["brand"]]
        df = df[df["brand"].isin(brands)]
    if parameters.get("channel"):
        channels = parameters["channel"] if isinstance(parameters["channel"], list) else [parameters["channel"]]
        df = df[df["channel"].isin(channels)]
    if parameters.get("start_date"):
        df = df[df["date"] >= pd.to_datetime(parameters["start_date"])]
    if parameters.get("end_date"):
        df = df[df["date"] <= pd.to_datetime(parameters["end_date"])]

    if df.empty:
        raise ValueError("No data matches the given filters.")
    return df


def generate_report(report_type: str, parameters: dict) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    df = _apply_filter(load_data(), parameters)

    custom_title = parameters.get("title")
    include = parameters.get("include_sheets")
    is_custom = report_type == "custom"
    no_charts = bool(parameters.get("no_charts"))

    if include is None:
        if is_custom:
            include = ["summary", "custom", "raw"]
            if parameters.get("comparison"):
                include.insert(1, "comparison")
        else:
            include = ["summary", "channel", "brand", "trends", "raw"]

    wb = Workbook()
    first_sheet = True

    def _get_ws(title, tab_color=NAVY):
        nonlocal first_sheet
        if first_sheet:
            ws = wb.active
            ws.title = title
            first_sheet = False
        else:
            ws = wb.create_sheet(title)
        ws.sheet_properties.tabColor = tab_color
        return ws

    if "summary" in include:
        _build_executive_summary(wb, df, _get_ws, custom_title, no_charts)
    if "channel" in include:
        _build_channel_performance(wb, df, _get_ws, no_charts)
    if "brand" in include:
        _build_brand_performance(wb, df, _get_ws, no_charts)
    if "trends" in include:
        _build_monthly_trends(wb, df, _get_ws, no_charts)
    if "comparison" in include and parameters.get("comparison"):
        _build_comparison_sheet(wb, load_data(), parameters, _get_ws)
    if "custom" in include and is_custom:
        _build_custom_sheet(wb, df, parameters, _get_ws, no_charts)
    if "raw" in include:
        _build_detailed_data(wb, df, _get_ws)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"campaign_report_{timestamp}.xlsx"
    wb.save(os.path.join(REPORTS_DIR, filename))
    return filename


# ═══════════════════════════════════════════════
# Sheet 1: Executive Summary
# ═══════════════════════════════════════════════

def _build_executive_summary(wb, df, _get_ws, custom_title=None, no_charts=False):
    ws = _get_ws("Executive Summary", NAVY)

    total_spend = df["spend_usd"].sum()
    total_revenue = df["revenue_usd"].sum()
    total_impressions = df["impressions"].sum()
    total_clicks = df["clicks"].sum()
    total_conversions = df["conversions"].sum()
    overall_roas = total_revenue / total_spend if total_spend else 0
    overall_ctr = total_clicks / total_impressions if total_impressions else 0
    overall_cpa = total_spend / total_conversions if total_conversions else 0
    overall_cpc = total_spend / total_clicks if total_clicks else 0
    profit = total_revenue - total_spend

    date_min = df["date"].min().strftime("%B %d, %Y")
    date_max = df["date"].max().strftime("%B %d, %Y")

    # Title block
    ws.merge_cells("B2:H2")
    ws["B2"] = custom_title or "MEDIA CAMPAIGN PERFORMANCE REPORT"
    ws["B2"].font = title_font
    ws["B2"].alignment = Alignment(horizontal="left")

    ws.merge_cells("B3:H3")
    ws["B3"] = f"{date_min}  —  {date_max}"
    ws["B3"].font = subtitle_font

    ws.merge_cells("B4:H4")
    brands_str = ", ".join(sorted(df["brand"].unique()))
    channels_str = ", ".join(sorted(df["channel"].unique()))
    ws["B4"] = f"Brands: {brands_str}  |  Channels: {channels_str}"
    ws["B4"].font = Font(color="888888", size=9, name="Calibri")

    # Separator
    for col in range(2, 9):
        ws.cell(row=5, column=col).border = Border(bottom=Side(style="medium", color=NAVY))

    # KPI cards row 1
    _kpi_card(ws, 7, 2, "Total Spend", f"${total_spend:,.0f}")
    _kpi_card(ws, 7, 4, "Total Revenue", f"${total_revenue:,.0f}")
    _kpi_card(ws, 7, 6, "Net Profit", f"${profit:,.0f}",
              "PROFITABLE" if profit > 0 else "LOSS")
    _kpi_card(ws, 7, 8, "Overall ROAS", f"{overall_roas:.2f}x")

    # KPI cards row 2
    _kpi_card(ws, 11, 2, "Impressions", f"{total_impressions:,.0f}")
    _kpi_card(ws, 11, 4, "Clicks", f"{total_clicks:,.0f}")
    _kpi_card(ws, 11, 6, "Conversions", f"{total_conversions:,.0f}")
    _kpi_card(ws, 11, 8, "Avg CTR", f"{overall_ctr * 100:.2f}%")

    # KPI cards row 3
    _kpi_card(ws, 15, 2, "Avg CPA", f"${overall_cpa:,.2f}")
    _kpi_card(ws, 15, 4, "Avg CPC", f"${overall_cpc:,.2f}")
    cpm = total_spend / total_impressions * 1000 if total_impressions else 0
    _kpi_card(ws, 15, 6, "Avg CPM", f"${cpm:,.2f}")
    _kpi_card(ws, 15, 8, "Total Campaigns", str(df["campaign_name"].nunique()))

    # Separator
    for col in range(2, 9):
        ws.cell(row=18, column=col).border = Border(bottom=Side(style="thin", color=MID_GRAY))

    # Top / Bottom performers mini-tables
    ws.cell(row=20, column=2, value="TOP PERFORMING CAMPAIGNS (by ROAS)").font = subheader_font

    camp_agg = df.groupby("campaign_name").agg(
        Spend=("spend_usd", "sum"),
        Revenue=("revenue_usd", "sum"),
        Conversions=("conversions", "sum"),
    ).reset_index()
    camp_agg["ROAS"] = camp_agg["Revenue"] / camp_agg["Spend"]
    camp_agg = camp_agg.sort_values("ROAS", ascending=False)

    top5 = camp_agg.head(5).rename(columns={"campaign_name": "Campaign"})
    _write_table(ws, top5, 21, 2,
                 currency_cols={"Spend", "Revenue"}, int_cols={"Conversions"}, ratio_cols={"ROAS"})

    ws.cell(row=28, column=2, value="LOWEST PERFORMING CAMPAIGNS (by ROAS)").font = subheader_font
    bottom5 = camp_agg.tail(5).sort_values("ROAS").rename(columns={"campaign_name": "Campaign"})
    _write_table(ws, bottom5, 29, 2,
                 currency_cols={"Spend", "Revenue"}, int_cols={"Conversions"}, ratio_cols={"ROAS"})

    if not no_charts:
        # Spend allocation pie chart
        spend_by_channel = df.groupby("channel")["spend_usd"].sum().reset_index()
        spend_by_channel.columns = ["Channel", "Spend"]

        chart_data_start = 37
        ws.cell(row=chart_data_start, column=2, value="Channel").font = header_font
        ws.cell(row=chart_data_start, column=2).fill = header_fill
        ws.cell(row=chart_data_start, column=3, value="Spend").font = header_font
        ws.cell(row=chart_data_start, column=3).fill = header_fill
        for i, (_, row) in enumerate(spend_by_channel.iterrows()):
            ws.cell(row=chart_data_start + 1 + i, column=2, value=row["Channel"])
            ws.cell(row=chart_data_start + 1 + i, column=3, value=row["Spend"])

        pie = PieChart()
        pie.title = "Spend Allocation by Channel"
        pie.style = 10
        data_ref = Reference(ws, min_col=3, min_row=chart_data_start, max_row=chart_data_start + len(spend_by_channel))
        cats_ref = Reference(ws, min_col=2, min_row=chart_data_start + 1, max_row=chart_data_start + len(spend_by_channel))
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.width = 16
        pie.height = 10

        labels = DataLabelList()
        labels.showPercent = True
        labels.showCatName = True
        labels.showVal = False
        pie.dataLabels = labels

        ws.add_chart(pie, "E20")

    ws.column_dimensions["A"].width = 3
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 20

    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════
# Sheet 2: Channel Performance
# ═══════════════════════════════════════════════

def _build_channel_performance(wb, df, _get_ws, no_charts=False):
    ws = _get_ws("Channel Performance", ACCENT_BLUE)

    ws.merge_cells("B2:I2")
    ws["B2"] = "CHANNEL PERFORMANCE BREAKDOWN"
    ws["B2"].font = Font(bold=True, color=NAVY, size=14, name="Calibri")

    ch = df.groupby("channel").agg(
        Impressions=("impressions", "sum"),
        Clicks=("clicks", "sum"),
        Spend=("spend_usd", "sum"),
        Conversions=("conversions", "sum"),
        Revenue=("revenue_usd", "sum"),
    ).reset_index()
    ch["CTR"] = ch["Clicks"] / ch["Impressions"]
    ch["CPC"] = ch["Spend"] / ch["Clicks"]
    ch["CPA"] = ch["Spend"] / ch["Conversions"]
    ch["ROAS"] = ch["Revenue"] / ch["Spend"]
    ch["Conv Rate"] = ch["Conversions"] / ch["Clicks"]
    ch = ch.rename(columns={"channel": "Channel"})
    ch = ch[["Channel", "Impressions", "Clicks", "CTR", "Spend", "Conversions", "Conv Rate", "Revenue", "ROAS", "CPC", "CPA"]]
    ch = ch.sort_values("ROAS", ascending=False)

    end_row = _write_table(ws, ch, 4, 2,
                           currency_cols={"Spend", "Revenue", "CPC", "CPA"},
                           pct_cols={"CTR", "Conv Rate"},
                           int_cols={"Impressions", "Clicks", "Conversions"},
                           ratio_cols={"ROAS"})

    if not no_charts:
        bar = BarChart()
        bar.type = "col"
        bar.title = "ROAS by Channel"
        bar.style = 10
        bar.y_axis.title = "ROAS"
        roas_col_idx = list(ch.columns).index("ROAS") + 2
        data_ref = Reference(ws, min_col=roas_col_idx, min_row=4, max_row=4 + len(ch))
        cats_ref = Reference(ws, min_col=2, min_row=5, max_row=4 + len(ch))
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.shape = 4
        bar.width = 18
        bar.height = 12
        ws.add_chart(bar, f"B{end_row + 2}")

        bar2 = BarChart()
        bar2.type = "col"
        bar2.title = "Total Spend by Channel"
        bar2.style = 10
        bar2.y_axis.title = "Spend ($)"
        bar2.y_axis.numFmt = '$#,##0'
        spend_col_idx = list(ch.columns).index("Spend") + 2
        data_ref2 = Reference(ws, min_col=spend_col_idx, min_row=4, max_row=4 + len(ch))
        bar2.add_data(data_ref2, titles_from_data=True)
        bar2.set_categories(cats_ref)
        bar2.shape = 4
        bar2.width = 18
        bar2.height = 12
        ws.add_chart(bar2, f"H{end_row + 2}")

    ws.column_dimensions["A"].width = 3
    _auto_width(ws, min_width=12)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════
# Sheet 3: Brand Performance
# ═══════════════════════════════════════════════

def _build_brand_performance(wb, df, _get_ws, no_charts=False):
    ws = _get_ws("Brand Performance", ACCENT_GREEN)

    ws.merge_cells("B2:I2")
    ws["B2"] = "BRAND PERFORMANCE COMPARISON"
    ws["B2"].font = Font(bold=True, color=NAVY, size=14, name="Calibri")

    br = df.groupby("brand").agg(
        Impressions=("impressions", "sum"),
        Clicks=("clicks", "sum"),
        Spend=("spend_usd", "sum"),
        Conversions=("conversions", "sum"),
        Revenue=("revenue_usd", "sum"),
    ).reset_index()
    br["CTR"] = br["Clicks"] / br["Impressions"]
    br["CPA"] = br["Spend"] / br["Conversions"]
    br["ROAS"] = br["Revenue"] / br["Spend"]
    br["Profit"] = br["Revenue"] - br["Spend"]
    br = br.rename(columns={"brand": "Brand"})
    br = br[["Brand", "Spend", "Revenue", "Profit", "ROAS", "Impressions", "Clicks", "CTR", "Conversions", "CPA"]]

    end_row = _write_table(ws, br, 4, 2,
                           currency_cols={"Spend", "Revenue", "Profit", "CPA"},
                           pct_cols={"CTR"},
                           int_cols={"Impressions", "Clicks", "Conversions"},
                           ratio_cols={"ROAS"})

    # Brand × Channel breakdown
    ws.cell(row=end_row + 2, column=2, value="BRAND × CHANNEL BREAKDOWN").font = subheader_font

    bc = df.groupby(["brand", "channel"]).agg(
        Spend=("spend_usd", "sum"),
        Revenue=("revenue_usd", "sum"),
        Conversions=("conversions", "sum"),
    ).reset_index()
    bc["ROAS"] = bc["Revenue"] / bc["Spend"]
    bc["CPA"] = bc["Spend"] / bc["Conversions"]
    bc = bc.rename(columns={"brand": "Brand", "channel": "Channel"})
    bc = bc.sort_values(["Brand", "ROAS"], ascending=[True, False])

    end_row2 = _write_table(ws, bc, end_row + 3, 2,
                            currency_cols={"Spend", "Revenue", "CPA"},
                            int_cols={"Conversions"},
                            ratio_cols={"ROAS"})

    if not no_charts:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Revenue & Spend by Brand"
        bar.style = 10
        bar.y_axis.numFmt = '$#,##0'
        spend_col = list(br.columns).index("Spend") + 2
        rev_col = list(br.columns).index("Revenue") + 2
        cats = Reference(ws, min_col=2, min_row=5, max_row=4 + len(br))
        d1 = Reference(ws, min_col=spend_col, min_row=4, max_row=4 + len(br))
        d2 = Reference(ws, min_col=rev_col, min_row=4, max_row=4 + len(br))
        bar.add_data(d1, titles_from_data=True)
        bar.add_data(d2, titles_from_data=True)
        bar.set_categories(cats)
        bar.width = 18
        bar.height = 12
        ws.add_chart(bar, f"H4")

    ws.column_dimensions["A"].width = 3
    _auto_width(ws, min_width=12)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════
# Sheet 4: Monthly Trends
# ═══════════════════════════════════════════════

def _build_monthly_trends(wb, df, _get_ws, no_charts=False):
    ws = _get_ws("Monthly Trends", "E67E22")

    ws.merge_cells("B2:I2")
    ws["B2"] = "MONTHLY PERFORMANCE TRENDS"
    ws["B2"].font = Font(bold=True, color=NAVY, size=14, name="Calibri")

    df_m = df.copy()
    df_m["month"] = df_m["date"].dt.to_period("M").astype(str)

    monthly = df_m.groupby("month").agg(
        Spend=("spend_usd", "sum"),
        Revenue=("revenue_usd", "sum"),
        Impressions=("impressions", "sum"),
        Clicks=("clicks", "sum"),
        Conversions=("conversions", "sum"),
    ).reset_index()
    monthly["ROAS"] = monthly["Revenue"] / monthly["Spend"]
    monthly["CTR"] = monthly["Clicks"] / monthly["Impressions"]
    monthly["CPA"] = monthly["Spend"] / monthly["Conversions"]
    monthly["Profit"] = monthly["Revenue"] - monthly["Spend"]
    monthly = monthly.rename(columns={"month": "Month"})
    monthly = monthly[["Month", "Spend", "Revenue", "Profit", "ROAS", "Impressions", "Clicks", "CTR", "Conversions", "CPA"]]

    end_row = _write_table(ws, monthly, 4, 2,
                           currency_cols={"Spend", "Revenue", "Profit", "CPA"},
                           pct_cols={"CTR"},
                           int_cols={"Impressions", "Clicks", "Conversions"},
                           ratio_cols={"ROAS"})

    # MoM % change
    if len(monthly) > 1:
        ws.cell(row=end_row + 1, column=2, value="MONTH-OVER-MONTH CHANGE (%)").font = subheader_font

        change_data = []
        for i in range(1, len(monthly)):
            row_data = {"Month": f"{monthly.iloc[i-1]['Month']} → {monthly.iloc[i]['Month']}"}
            for col in ["Spend", "Revenue", "Impressions", "Clicks", "Conversions", "ROAS"]:
                prev = monthly.iloc[i-1][col]
                curr = monthly.iloc[i][col]
                pct = (curr - prev) / prev if prev else 0
                row_data[f"{col} Δ%"] = pct
            change_data.append(row_data)
        change_df = pd.DataFrame(change_data)
        end_row2 = _write_table(ws, change_df, end_row + 2, 2,
                                pct_cols={c for c in change_df.columns if "Δ%" in c})

    # Weekly granularity table — placed below the MoM change table
    df_w = df.copy()
    df_w["week"] = df_w["date"].dt.isocalendar().week.astype(int)
    df_w["year"] = df_w["date"].dt.year
    weekly = df_w.groupby(["year", "week"]).agg(
        Spend=("spend_usd", "sum"),
        Revenue=("revenue_usd", "sum"),
        Conversions=("conversions", "sum"),
    ).reset_index()
    weekly["ROAS"] = weekly["Revenue"] / weekly["Spend"]
    weekly["Week"] = weekly.apply(lambda r: f"W{int(r['week']):02d}", axis=1)
    weekly = weekly[["Week", "Spend", "Revenue", "Conversions", "ROAS"]]

    weekly_start = (end_row2 + 2) if len(monthly) > 1 else (end_row + 2)
    ws.cell(row=weekly_start, column=2, value="WEEKLY PERFORMANCE DETAIL").font = subheader_font
    weekly_end = _write_table(ws, weekly, weekly_start + 1, 2,
                              currency_cols={"Spend", "Revenue"},
                              int_cols={"Conversions"},
                              ratio_cols={"ROAS"})

    if not no_charts:
        line = LineChart()
        line.title = "Revenue & Spend Trend"
        line.style = 10
        line.y_axis.numFmt = '$#,##0'
        line.width = 18
        line.height = 12

        spend_col = list(monthly.columns).index("Spend") + 2
        rev_col = list(monthly.columns).index("Revenue") + 2
        cats = Reference(ws, min_col=2, min_row=5, max_row=4 + len(monthly))
        d1 = Reference(ws, min_col=spend_col, min_row=4, max_row=4 + len(monthly))
        d2 = Reference(ws, min_col=rev_col, min_row=4, max_row=4 + len(monthly))
        line.add_data(d1, titles_from_data=True)
        line.add_data(d2, titles_from_data=True)
        line.set_categories(cats)
        ws.add_chart(line, f"B{weekly_end + 2}")

    ws.column_dimensions["A"].width = 3
    _auto_width(ws, min_width=12)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════
# Sheet 5: Detailed Data
# ═══════════════════════════════════════════════

def _build_detailed_data(wb, df, _get_ws):
    ws = _get_ws("Raw Data", MID_GRAY)

    detail = df.sort_values(["date", "brand", "channel"]).copy()
    detail["date"] = detail["date"].dt.strftime("%Y-%m-%d")
    headers = list(detail.columns)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    currency_cols = {"spend_usd", "revenue_usd", "cpa_usd", "cpc_usd", "cpm_usd"}
    pct_cols = {"ctr", "conversion_rate"}

    for ri, row in enumerate(detail.itertuples(index=False), 2):
        is_stripe = ri % 2 == 0
        for ci, (col_name, val) in enumerate(zip(headers, row), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = body_font
            cell.border = thin_border
            if is_stripe:
                cell.fill = stripe_fill
            if col_name in currency_cols:
                cell.number_format = '$#,##0.00'
            elif col_name in pct_cols:
                cell.number_format = '0.00%'
            elif col_name == "roas":
                cell.number_format = '0.00'
            elif isinstance(val, (int, float)):
                cell.number_format = '#,##0'

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(detail) + 1}"
    ws.freeze_panes = "A2"
    _auto_width(ws)


# ═══════════════════════════════════════════════
# Sheet: Period Comparison (custom)
# ═══════════════════════════════════════════════

def _build_comparison_sheet(wb, raw_df, parameters, _get_ws):
    ws = _get_ws("Period Comparison", "8E44AD")

    comp = parameters["comparison"]
    p1_start = pd.to_datetime(comp["period_1_start"])
    p1_end = pd.to_datetime(comp["period_1_end"])
    p2_start = pd.to_datetime(comp["period_2_start"])
    p2_end = pd.to_datetime(comp["period_2_end"])
    p1_label = comp.get("period_1_label", f"{p1_start.date()} to {p1_end.date()}")
    p2_label = comp.get("period_2_label", f"{p2_start.date()} to {p2_end.date()}")

    df = raw_df.copy()
    df["date"] = pd.to_datetime(df["date"])

    # Apply brand/channel filters
    if parameters.get("brand"):
        brands = parameters["brand"] if isinstance(parameters["brand"], list) else [parameters["brand"]]
        df = df[df["brand"].isin(brands)]
    if parameters.get("channel"):
        channels = parameters["channel"] if isinstance(parameters["channel"], list) else [parameters["channel"]]
        df = df[df["channel"].isin(channels)]

    df1 = df[(df["date"] >= p1_start) & (df["date"] <= p1_end)]
    df2 = df[(df["date"] >= p2_start) & (df["date"] <= p2_end)]

    ws.merge_cells("B2:J2")
    ws["B2"] = f"PERIOD COMPARISON: {p1_label} vs {p2_label}"
    ws["B2"].font = Font(bold=True, color=NAVY, size=14, name="Calibri")

    def _agg(d):
        return {
            "Spend": d["spend_usd"].sum(),
            "Revenue": d["revenue_usd"].sum(),
            "Impressions": d["impressions"].sum(),
            "Clicks": d["clicks"].sum(),
            "Conversions": d["conversions"].sum(),
            "ROAS": d["revenue_usd"].sum() / d["spend_usd"].sum() if d["spend_usd"].sum() else 0,
            "CTR": d["clicks"].sum() / d["impressions"].sum() if d["impressions"].sum() else 0,
            "CPA": d["spend_usd"].sum() / d["conversions"].sum() if d["conversions"].sum() else 0,
        }

    # Overall comparison
    a1, a2 = _agg(df1), _agg(df2)
    rows = []
    for metric in a1:
        v1, v2 = a1[metric], a2[metric]
        change = (v2 - v1) / v1 if v1 else 0
        rows.append({"Metric": metric, p1_label: v1, p2_label: v2, "Δ Change": v2 - v1, "Δ %": change})
    comp_df = pd.DataFrame(rows)

    ws.cell(row=4, column=2, value="OVERALL COMPARISON").font = subheader_font
    end_row = _write_table(ws, comp_df, 5, 2,
                           currency_cols={p1_label, p2_label, "Δ Change"},
                           pct_cols={"Δ %"})

    # By-channel comparison
    group_col = "channel"
    if len(df["channel"].unique()) == 1 and len(df["brand"].unique()) > 1:
        group_col = "brand"

    ws.cell(row=end_row + 2, column=2,
            value=f"BY {group_col.upper()} COMPARISON").font = subheader_font

    chan_rows = []
    for name in sorted(df[group_col].unique()):
        s1 = df1[df1[group_col] == name]
        s2 = df2[df2[group_col] == name]
        spend1, spend2 = s1["spend_usd"].sum(), s2["spend_usd"].sum()
        rev1, rev2 = s1["revenue_usd"].sum(), s2["revenue_usd"].sum()
        roas1 = rev1 / spend1 if spend1 else 0
        roas2 = rev2 / spend2 if spend2 else 0
        chan_rows.append({
            group_col.title(): name,
            f"{p1_label} Spend": spend1, f"{p2_label} Spend": spend2,
            f"{p1_label} Revenue": rev1, f"{p2_label} Revenue": rev2,
            f"{p1_label} ROAS": roas1, f"{p2_label} ROAS": roas2,
        })
    chan_df = pd.DataFrame(chan_rows)
    spend_cols = {c for c in chan_df.columns if "Spend" in c or "Revenue" in c}
    roas_cols = {c for c in chan_df.columns if "ROAS" in c}
    _write_table(ws, chan_df, end_row + 3, 2, currency_cols=spend_cols, ratio_cols=roas_cols)

    ws.column_dimensions["A"].width = 3
    _auto_width(ws, min_width=12)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════
# Sheet: Custom Grouped Data (custom)
# ═══════════════════════════════════════════════

METRIC_MAP = {
    "spend": ("spend_usd", "sum", "Spend", "currency"),
    "revenue": ("revenue_usd", "sum", "Revenue", "currency"),
    "impressions": ("impressions", "sum", "Impressions", "int"),
    "clicks": ("clicks", "sum", "Clicks", "int"),
    "conversions": ("conversions", "sum", "Conversions", "int"),
}

DERIVED_METRICS = {"roas", "ctr", "cpa", "cpc", "cpm", "profit"}


def _build_custom_sheet(wb, df, parameters, _get_ws, no_charts=False):
    ws = _get_ws("Custom Analysis", DARK_BLUE)

    group_by = parameters.get("group_by", "brand")
    requested_metrics = parameters.get("metrics")
    sort_by = parameters.get("sort_by")
    sort_order = parameters.get("sort_order", "desc")
    top_n = parameters.get("top_n")

    group_col_map = {
        "brand": "brand",
        "channel": "channel",
        "campaign": "campaign_name",
        "month": "month",
        "week": "week",
    }

    if group_by in ("month", "week"):
        if group_by == "month":
            df["month"] = df["date"].dt.to_period("M").astype(str)
        else:
            df["week"] = "W" + df["date"].dt.isocalendar().week.astype(int).astype(str).str.zfill(2)

    actual_col = group_col_map.get(group_by, "brand")

    agg = df.groupby(actual_col).agg(
        Spend=("spend_usd", "sum"),
        Revenue=("revenue_usd", "sum"),
        Impressions=("impressions", "sum"),
        Clicks=("clicks", "sum"),
        Conversions=("conversions", "sum"),
    ).reset_index()

    # Compute derived metrics
    agg["ROAS"] = agg["Revenue"] / agg["Spend"].replace(0, float("nan"))
    agg["CTR"] = agg["Clicks"] / agg["Impressions"].replace(0, float("nan"))
    agg["CPA"] = agg["Spend"] / agg["Conversions"].replace(0, float("nan"))
    agg["CPC"] = agg["Spend"] / agg["Clicks"].replace(0, float("nan"))
    agg["CPM"] = (agg["Spend"] / agg["Impressions"].replace(0, float("nan"))) * 1000
    agg["Profit"] = agg["Revenue"] - agg["Spend"]
    agg = agg.fillna(0)

    agg = agg.rename(columns={actual_col: group_by.title()})

    # Select metrics
    all_metric_cols = ["Spend", "Revenue", "Profit", "ROAS", "Impressions", "Clicks",
                       "CTR", "Conversions", "CPA", "CPC", "CPM"]
    if requested_metrics:
        name_map = {
            "spend": "Spend", "revenue": "Revenue", "profit": "Profit",
            "roas": "ROAS", "impressions": "Impressions", "clicks": "Clicks",
            "ctr": "CTR", "conversions": "Conversions", "cpa": "CPA",
            "cpc": "CPC", "cpm": "CPM",
        }
        selected = [name_map[m] for m in requested_metrics if m in name_map]
        if not selected:
            selected = all_metric_cols
    else:
        selected = all_metric_cols

    keep_cols = [group_by.title()] + [c for c in selected if c in agg.columns]
    agg = agg[keep_cols]

    # Sort
    sort_name_map = {
        "roas": "ROAS", "spend": "Spend", "revenue": "Revenue",
        "conversions": "Conversions", "cpa": "CPA", "clicks": "Clicks",
    }
    if sort_by and sort_name_map.get(sort_by) in agg.columns:
        agg = agg.sort_values(sort_name_map[sort_by], ascending=(sort_order == "asc"))

    # Top N
    if top_n and isinstance(top_n, int) and top_n > 0:
        agg = agg.head(top_n)

    # Determine column types
    currency_cols = {c for c in agg.columns if c in ("Spend", "Revenue", "Profit", "CPA", "CPC", "CPM")}
    pct_cols = {c for c in agg.columns if c in ("CTR",)}
    ratio_cols = {c for c in agg.columns if c in ("ROAS",)}
    int_cols = {c for c in agg.columns if c in ("Impressions", "Clicks", "Conversions")}

    title_text = parameters.get("title") or f"Custom Analysis — Grouped by {group_by.title()}"
    ws.merge_cells("B2:J2")
    ws["B2"] = title_text.upper()
    ws["B2"].font = Font(bold=True, color=NAVY, size=14, name="Calibri")

    filters_desc = []
    if parameters.get("brand"):
        b = parameters["brand"]
        filters_desc.append(f"Brands: {b if isinstance(b, str) else ', '.join(b)}")
    if parameters.get("channel"):
        c = parameters["channel"]
        filters_desc.append(f"Channels: {c if isinstance(c, str) else ', '.join(c)}")
    if parameters.get("start_date") or parameters.get("end_date"):
        filters_desc.append(f"Date range: {parameters.get('start_date', 'start')} to {parameters.get('end_date', 'end')}")
    if top_n:
        filters_desc.append(f"Top {top_n} by {sort_by or 'default'}")

    if filters_desc:
        ws.merge_cells("B3:J3")
        ws["B3"] = " | ".join(filters_desc)
        ws["B3"].font = Font(color="888888", size=9, name="Calibri")

    end_row = _write_table(ws, agg, 5, 2,
                           currency_cols=currency_cols,
                           pct_cols=pct_cols,
                           ratio_cols=ratio_cols,
                           int_cols=int_cols)

    # Add a bar chart for the first numeric metric
    if not no_charts and len(agg) > 1 and len(selected) > 0:
        chart_metric = selected[0]
        chart_col_idx = list(agg.columns).index(chart_metric) + 2
        bar = BarChart()
        bar.type = "col"
        bar.title = f"{chart_metric} by {group_by.title()}"
        bar.style = 10
        if chart_metric in currency_cols:
            bar.y_axis.numFmt = '$#,##0'
        data_ref = Reference(ws, min_col=chart_col_idx, min_row=5, max_row=5 + len(agg))
        cats_ref = Reference(ws, min_col=2, min_row=6, max_row=5 + len(agg))
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width = 18
        bar.height = 12
        ws.add_chart(bar, f"B{end_row + 2}")

    ws.column_dimensions["A"].width = 3
    _auto_width(ws, min_width=12)
    ws.sheet_view.showGridLines = False
